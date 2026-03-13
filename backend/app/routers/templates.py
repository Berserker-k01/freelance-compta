from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from ..database import get_db
from ..models import Company, Account, EntryLine, Entry, Journal
from ..models_user import User
from ..auth_utils import get_current_user
from ..services.injector import ExcelInjector
import os
from datetime import datetime
from typing import Optional

router = APIRouter(
    prefix="/templates",
    tags=["templates"],
    responses={404: {"description": "Not found"}},
)

# Path resolution: this file is at backend/app/routers/templates.py
# BASE_DIR → backend/
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUT_DIR = os.path.join(BASE_DIR, "temp_exports")

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

# ---------------------------------------------------------------------------
# TEMPLATE CRUD ENDPOINTS (BEFORE dynamic routes)
# ---------------------------------------------------------------------------

from .. import models
import json
from fastapi import UploadFile, File, Form
from openpyxl.utils import get_column_letter

@router.get("/list")
def list_templates(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """List all available report templates."""
    templates = db.query(models.ReportTemplate).filter(models.ReportTemplate.user_id == current_user.id).all()
    return templates

@router.post("/upload")
async def upload_dynamic_template(
    file: UploadFile = File(...), 
    name: str = Form("Nouveau Canevas"), 
    year: int = Form(2026), 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Smart Loader: Upload an Excel template and auto-map basic tags."""
    import openpyxl
    file_path = os.path.join(BASE_DIR, "templates", f"dynamic_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
    with open(file_path, "wb") as f:
        f.write(await file.read())

    # Auto-mapping
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    mapping = {}
    
    # Auto-mapping Expert : Base de Connaissances exhaustive SYSCOHADA/OTR
    # Format: "mot clé le plus long d'abord": [(décalage_col, "règle")]
    # En SYSCOHADA, le Bilan Actif place souvent ses titres en B, BRUT en E (+3), AMORT en F (+4).
    # Le Bilan Passif place ses titres en B, NET en F (+4).
    # L'ordre du dictionnaire est respecté: le plus long match et on sort de la boucle !
    KNOWLEDGE_BASE = {
        "brevets, licences,logiciels et droits similaires": [(3, "20*"), (4, "ABS(280*)")],
        "banques, chèques postaux, caisse et assimilés": [(3, "SD(52*, 53*, 57*)")],
        "aménagements, agencements et installations": [(3, "232*, 233*, 241*, 242*"), (4, "ABS(283*, 284*)")],
        "matériel, mobilier et actifs biologiques": [(3, "24*"), (4, "ABS(284*)")],
        "frais de développement et de prospection": [(3, "20*"), (4, "ABS(280*)")],
        "avances et acomptes versés sur immobilisations": [(3, "26*")],
        "matières premières et fournitures liées": [(3, "32*"), (4, "ABS(39*)")],
        "emprunts et dettes financières diverses": [(4, "SC(16*, 17*)")],
        "fournisseurs et comptes rattachés": [(4, "SC(401*, 402*, 408*)")],
        "variation de stocks de matières premières": [(1, "6032*")],
        "fonds commercial et droit au bail": [(3, "20*"), (4, "ABS(280*)")],
        "provisions pour dépréciation des stocks": [(4, "ABS(39*)")],
        "variation de stocks de marchandises": [(1, "6031*")],
        "numéro d'identification fiscale": [(1, "#nif")],
        "titres de participation": [(3, "27*")],
        "autres immobilisations financières": [(3, "27*")],
        "produits intermédiaires": [(3, "37*"), (4, "ABS(39*)")],
        "clients et comptes rattachés": [(3, "SD(41*)"), (4, "ABS(49*)")],
        "primes liées au capital social": [(4, "SC(105*)")],
        "subventions d'investissement": [(4, "SC(14*)")],
        "résultat net de l'exercice": [(4, "SC(13*)")], # Note: Si perte, ça peut afficher 0 avec SC(). Généralement, Bilan Passif = SC pour Bénéfice, et on gère les pertes en -SD(13*). L'OTR affiche souvent le compte 13 "de son sens". Testons avec -13* pour l'instant.
        "dettes fiscales et sociales": [(4, "SC(42*, 43*, 44*)")],
        "transferts de charges d'exploitation": [(1, "-781*")],
        "ventes de produits fabriqués": [(1, "-702*, -703*, -704*")],
        "dotations aux amortissements": [(1, "68*")],
        "participation des travailleurs": [(1, "87*")],
        "autres approvisionnements": [(3, "33*"), (4, "ABS(39*)")],
        "valeurs à encaisser": [(3, "SD(51*)")],
        "avances et acomptes reçus": [(4, "SC(419*)")],
        "provisions réglementées": [(4, "SC(15*)")],
        "réserves indisponibles": [(4, "SC(111*, 112*)")],
        "banques, découverts": [(4, "SC(52*, 53*, 56*)")],
        "ventes de marchandises": [(1, "-701*")],
        "travaux, services vendus": [(1, "-705*, -706*")],
        "subventions d'exploitation": [(1, "-74*")],
        "achats de matières premières": [(1, "602*")],
        "charges non justifiées": [(1, "658*")],
        "apporteurs capital non appelé": [(4, "109*")],
        "produits en cours": [(3, "34*"), (4, "ABS(39*)")],
        "services en cours": [(3, "35*"), (4, "ABS(39*)")],
        "produits accessoires": [(1, "-707*")],
        "achats de marchandises": [(1, "601*")],
        "reprises de provisions": [(1, "-78*")],
        "impôt sur le résultat": [(1, "89*")],
        "amendes et pénalités": [(1, "657*")],
        "dons et libéralités": [(1, "6234*")],
        "produits finis": [(3, "36*"), (4, "ABS(39*)")],
        "immobilisations corporelles": [(3, "21*, 22*, 23*, 24*"), (4, "ABS(281*, 282*, 283*, 284*)")],
        "immobilisations incorporelles": [(3, "20*"), (4, "ABS(280*)")],
        "autres créances": [(3, "SD(409*, 44*, 45*, 46*, 47*, 48*)")],
        "autres dettes": [(4, "SC(45*, 46*, 47*, 48*)")],
        "report à nouveau": [(4, "-12*")], # Laisse avec - car le report peut être débiteur ou créditeur et on veut la valeur mathématique
        "réserves libres": [(4, "SC(118*)")],
        "chiffre d'affaires": [(1, "-70*")],
        "autres produits": [(1, "-75*")],
        "services extérieurs": [(1, "62*, 63*")],
        "frais de personnel": [(1, "66*")],
        "produits financiers": [(1, "-77*")],
        "charges financières": [(1, "67*")],
        "nom du dirigeant": [(1, "#dirigeant_nom")],
        "effectif total brut": [(1, "#effectif_hommes")],
        "marchandises": [(3, "31*"), (4, "ABS(39*)")],
        "bâtiments": [(3, "23*"), (4, "ABS(283*)")],
        "impôts et taxes": [(1, "64*")],
        "autres charges": [(1, "65*")],
        "terrains": [(3, "22*"), (4, "ABS(282*)")],
        "banques": [(3, "SD(52*, 53*, 57*)")],
        "capital": [(4, "SC(101*, 102*)")],
        "nif": [(1, "#nif")]
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            for row in ws.iter_rows(min_row=1, max_row=150, min_col=1, max_col=10):
                for cell in row:
                    val = str(cell.value).strip().lower() if cell.value else ""
                    if len(val) > 3:
                        for keyword, rules in KNOWLEDGE_BASE.items():
                            if keyword in val:
                                for col_offset, account_rule in rules:
                                    target_col_idx = cell.column + col_offset
                                    cell_addr = f"{sheet_name}!{get_column_letter(target_col_idx)}{cell.row}"
                                    mapping[cell_addr] = account_rule
                                break # Match trouvé => on ne cherche plus d'autres mots pour cette cellule
        except Exception:
            continue
    wb.close()

    new_template = models.ReportTemplate(
        name=name,
        year=year,
        file_path=file_path,
        mapping_config=json.dumps(mapping),
        user_id=current_user.id
    )
    db.add(new_template)
    db.commit()
    db.refresh(new_template)
    return {"message": "Template importé et auto-mappé avec succès.", "template": new_template}


# ---------------------------------------------------------------------------
# PREREQUISITE VALIDATION ENDPOINT
# ---------------------------------------------------------------------------

# Account class prefixes that MUST be populated for a meaningful liasse
# { "label": ("prefix1", "prefix2", ...) }
REQUIRED_ACCOUNT_CLASSES = {
    "Comptes de Capitaux (Classe 1 — Capital, Réserves, Emprunts)": ("1",),
    "Stocks (Classe 3)": ("3",),
    "Comptes de Tiers (Classe 4 — Clients, Fournisseurs)": ("4",),
    "Trésorerie (Classe 5 — Banque, Caisse)": ("52", "53", "57"),
    "Charges d'Exploitation (Classe 6)": ("6",),
    "Produits d'Exploitation (Classe 7 — CA)": ("7",),
}


@router.get("/validate/{company_id}")
def validate_prerequisites(
    company_id: str, 
    document_id: Optional[str] = None, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Vérifie tous les prérequis avant de générer la liasse.
    Retourne :
      {
        "ready": bool,
        "blockers": [str],     # erreurs bloquantes
        "warnings": [str],     # avertissements non bloquants
        "checks": [            # détail de chaque contrôle
          { "name": str, "status": "OK"|"WARNING"|"KO", "detail": str }
        ]
      }
    """
    checks = []
    blockers = []
    warnings = []

    # ------------------------------------------------------------------ #
    # 1. La société existe et appartient à l'utilisateur                 #
    # ------------------------------------------------------------------ #
    company = db.query(Company).filter(Company.id == company_id, Company.user_id == current_user.id).first()
    if not company:
        return {
            "ready": False,
            "blockers": ["Société introuvable (ID invalide)."],
            "warnings": [],
            "checks": [{"name": "Société", "status": "KO", "detail": "Société introuvable."}],
        }
    checks.append({"name": "Société", "status": "OK", "detail": f"Dossier : {company.name}"})

    # ------------------------------------------------------------------ #
    # 2. Le fichier template Excel existe (Vérification repoussée à la génération)
    # ------------------------------------------------------------------ #
    checks.append({
        "name": "Modèle Excel",
        "status": "OK",
        "detail": "Vérification effectuée lors de la sélection du modèle dynamique.",
    })

    # ------------------------------------------------------------------ #
    # 3. Des écritures comptables existent                                 #
    # ------------------------------------------------------------------ #
    entry_query = (
        db.query(func.count(Entry.id))
        .join(Journal, Journal.id == Entry.journal_id)
        .filter(Journal.company_id == company_id)
    )
    if document_id:
        entry_query = entry_query.filter(Entry.document_id == document_id)

    nb_entries = entry_query.scalar() or 0

    if nb_entries == 0:
        source_label = f"le document #{document_id}" if document_id else "ce dossier"
        msg = (
            f"Aucune écriture comptable trouvée pour {source_label}. "
            "Veuillez d'abord importer une balance générale."
        )
        checks.append({"name": "Écritures Comptables", "status": "KO", "detail": msg})
        blockers.append(msg)
    else:
        checks.append({
            "name": "Écritures Comptables",
            "status": "OK",
            "detail": f"{nb_entries:,} écriture(s) disponible(s) pour la génération.",
        })

    # ------------------------------------------------------------------ #
    # 4. Balance équilibrée (Σ Débit = Σ Crédit)                         #
    # ------------------------------------------------------------------ #
    if nb_entries > 0:
        q_base = (
            db.query(
                func.sum(EntryLine.debit).label("total_debit"),
                func.sum(EntryLine.credit).label("total_credit"),
            )
            .join(Entry, Entry.id == EntryLine.entry_id)
            .join(Journal, Journal.id == Entry.journal_id)
            .filter(Journal.company_id == company_id)
        )
        if document_id:
            q_base = q_base.filter(Entry.document_id == document_id)

        row = q_base.first()
        total_debit  = float(row.total_debit or 0)
        total_credit = float(row.total_credit or 0)
        diff = round(abs(total_debit - total_credit), 2)

        if diff > 0.01:
            msg = (
                f"La balance n'est pas équilibrée : Σ Débit = {total_debit:,.2f} / "
                f"Σ Crédit = {total_credit:,.2f} — Écart : {diff:,.2f} FCFA. "
                "La liasse sera générée mais les totaux pourraient être faux."
            )
            checks.append({"name": "Équilibre de la Balance", "status": "WARNING", "detail": msg})
            warnings.append(msg)
        else:
            checks.append({
                "name": "Équilibre de la Balance",
                "status": "OK",
                "detail": f"Balance équilibrée — Σ D = Σ C = {total_debit:,.2f} FCFA.",
            })

    # ------------------------------------------------------------------ #
    # 5. Comptes clés alimentés (Classes requises)                        #
    # ------------------------------------------------------------------ #
    if nb_entries > 0:
        # Get all account codes that have movements for this company
        populated_codes = {
            row.code
            for row in (
                db.query(Account.code)
                .join(EntryLine, EntryLine.account_id == Account.id)
                .join(Entry, Entry.id == EntryLine.entry_id)
                .join(Journal, Journal.id == Entry.journal_id)
                .filter(Journal.company_id == company_id)
                .distinct()
                .all()
            )
        }

        for label, prefixes in REQUIRED_ACCOUNT_CLASSES.items():
            has_accounts = any(
                code.startswith(p) for code in populated_codes for p in prefixes
            )
            if not has_accounts:
                detail = (
                    f"Aucun mouvement trouvé pour les comptes commençant par "
                    f"{' / '.join(prefixes)}. "
                    "Les postes correspondants dans la liasse seront vides."
                )
                checks.append({"name": label, "status": "WARNING", "detail": detail})
                warnings.append(f"{label} : {detail}")
            else:
                # Find sample accounts
                samples = [c for c in sorted(populated_codes) if any(c.startswith(p) for p in prefixes)][:3]
                checks.append({
                    "name": label,
                    "status": "OK",
                    "detail": f"Comptes alimentés ex. : {', '.join(samples)}",
                })

    # ------------------------------------------------------------------ #
    # SUMMARY                                                              #
    # ------------------------------------------------------------------ #
    ready = len(blockers) == 0
    return {
        "ready": ready,
        "blockers": blockers,
        "warnings": warnings,
        "checks": checks,
    }





# ---------------------------------------------------------------------------
# GENERATE ENDPOINTS & PRE-FLIGHT
# ---------------------------------------------------------------------------

@router.get("/preflight/{company_id}")
def preflight_check(
    company_id: str, 
    document_id: Optional[str] = None, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Simule l'état Bilan/Résultat pour identifier un déséquilibre avant génération (Module 6).
    """
    from ..services.injector import ExcelInjector

    company = db.query(Company).filter(Company.id == company_id, Company.user_id == current_user.id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Société introuvable")

    injector = ExcelInjector(db, company_id, document_id=document_id)
    return injector.pre_flight_check()

@router.get("/generate/{company_id}")
async def generate_liasse(
    company_id: str,
    document_id: Optional[str] = None,
    template_id: Optional[str] = None,
    use_ia: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Generate the fiscal liasse (OTR/SYSCOHADA) by injecting account balances into the template."""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Société introuvable")

    if not template_id:
        raise HTTPException(status_code=400, detail="Veuillez sélectionner un modèle de déclaration.")

    tmpl = db.query(models.ReportTemplate).filter(
        models.ReportTemplate.id == template_id, 
        models.ReportTemplate.user_id == company.user_id
    ).first()
    if not tmpl or not tmpl.file_path or not os.path.exists(tmpl.file_path):
        raise HTTPException(status_code=404, detail="Modèle introuvable ou fichier source manquant sur le serveur.")

    working_template_path = tmpl.file_path
    template_prefix = tmpl.name
    try:
        mapping_config_dict = json.loads(tmpl.mapping_config)
    except Exception:
        mapping_config_dict = {}

    injector = ExcelInjector(db, company_id, document_id=document_id)

    # ---------------------------------------------------------
    # CHECKPOINT IA : QWEN 2.5 (Embarqué in-process)
    # ---------------------------------------------------------
    if use_ia:
        # Check plan permissions
        if not (current_user.is_superuser or (current_user.plan and current_user.plan.has_ai_access)):
            raise HTTPException(
                status_code=403, 
                detail="Votre abonnement ne permet pas l'Audit IA. Veuillez passer au forfait Premium."
            )
        
        try:
            from ..services.llm_audit import valider_coherence_ia_async
            
            # Preparation of the summary for the LLM
            preflight_data = injector.pre_flight_check()
            
            # Appeler l'IA local en mode asynchrone (non-bloquant FastAPI)
            audit_result = await valider_coherence_ia_async(preflight_data)
            
            if not audit_result.statut_conforme and audit_result.erreurs_bloquantes:
                errors_str = " | ".join(audit_result.erreurs_bloquantes)
                raise HTTPException(
                    status_code=422,
                    detail=f"Bloqué par l'IA d'Audit : {errors_str}"
                )
                
        except HTTPException:
            raise
        except Exception as e:
            # Fallback (ex: module introuvable / gguf absent)
            print(f"[IA Checkpoint] Ignoré. Erreur IA Embarquée : {e}")

    safe_name = company.name.replace(" ", "_").replace("/", "-")
    exercice = datetime.now().year
    
    # Preserver l'extension originale (.xlsx ou .xlsm) pour garder les macros
    ext = os.path.splitext(working_template_path)[1]
    if not ext:
        ext = ".xlsx"
        
    output_filename = f"Liasse_{template_prefix}_{safe_name}_{exercice}{ext}"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    try:
        injector.generate_report(
            template_path=working_template_path,
            output_path=output_path,
            mapping_config=mapping_config_dict,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        print(f"[generate_liasse] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur génération liasse : {str(e)}")

    return FileResponse(
        output_path,
        filename=output_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



# ---------------------------------------------------------------------------
# TEMPLATE CRUD BY ID (MUST be AFTER static routes)
# ---------------------------------------------------------------------------

@router.get("/{template_id}")
def get_template_by_id(template_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Get a single template by ID."""
    tmpl = db.query(models.ReportTemplate).filter(
        models.ReportTemplate.id == template_id,
        models.ReportTemplate.user_id == current_user.id
    ).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Modèle introuvable")
    return tmpl


@router.put("/{template_id}/mapping")
def update_template_mapping(template_id: str, payload: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Update the mapping configuration of a template."""
    tmpl = db.query(models.ReportTemplate).filter(
        models.ReportTemplate.id == template_id,
        models.ReportTemplate.user_id == current_user.id
    ).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Modèle introuvable")
    tmpl.mapping_config = payload.get("mapping_config", tmpl.mapping_config)
    db.commit()
    db.refresh(tmpl)
    return tmpl

@router.delete("/{template_id}")
def delete_template(template_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Delete a report template and its associated file."""
    tmpl = db.query(models.ReportTemplate).filter(
        models.ReportTemplate.id == template_id,
        models.ReportTemplate.user_id == current_user.id
    ).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Modèle introuvable")
    
    # Delete file if exists
    if tmpl.file_path and os.path.exists(tmpl.file_path):
        try:
            os.remove(tmpl.file_path)
        except Exception as e:
            print(f"Error deleting file {tmpl.file_path}: {e}")

    db.delete(tmpl)
    db.commit()
    return {"message": "Modèle supprimé avec succès"}
