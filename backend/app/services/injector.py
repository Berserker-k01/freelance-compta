import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func
from app import models
from app.services.injector_controls import compute_post_injection_controls
import os
import json
import re


class ExcelInjector:
    """
    Engine for injecting account balances from the database
    into a structured Excel template (SYSCOHADA / OTR format).

    Mapping rule syntax (supports multi-pattern, negation, ABS):
        "20*"           → Σ (Debit - Credit) for accounts starting with "20"
        "-70*"          → negate result (revenues are credit-heavy → flip positive)
        "ABS(280*)"     → absolute value (amortissements are credit → force positive)
        "601*, -603*"   → sum multiple patterns (achats + variation de stocks)
    """

    def __init__(
        self,
        db: Session,
        company_id: str,
        document_id: str | None = None,
        document_id_n1: str | None = None,
    ):
        self.db = db
        self.company_id = company_id
        self.document_id = document_id
        self.document_id_n1 = document_id_n1
        # { account_code: (debit_total, credit_total) }
        self.raw: dict[str, tuple[float, float]] = {}
        # { account_code: net_balance (debit - credit) }
        self.balances: dict[str, float] = {}
        self.raw_n1: dict[str, tuple[float, float]] = {}
        self.balances_n1: dict[str, float] = {}
        self.annexe_data: dict[str, str] = {}
        self._log: list[str] = []
        self.trace: list[dict] = []
        self.mapping_warnings: list[str] = []
        self.mapping_errors: list[str] = []

    @staticmethod
    def _normalize_code(code: str) -> str:
        # Keep only digits from account code labels such as "411.0" or "401-001"
        return re.sub(r"\D", "", str(code or ""))

    @staticmethod
    def _is_numeric_expr(rule: str) -> bool:
        r = (rule or "").strip()
        if not r:
            return False
        return not r.startswith("#")

    @staticmethod
    def parse_mapping_value(spec) -> tuple[str, str]:
        """
        Normalise une entrée de mapping.
        - Chaîne → règle pour l'exercice N (balance `document_id` ou cumul dossier).
        - {"rule": "...", "period": "n"|"n1"} → exercice courant ou N-1.
        """
        if isinstance(spec, dict):
            rule = str(spec.get("rule", "")).strip()
            period = str(spec.get("period", "n")).strip().lower()
            if period not in ("n", "n1"):
                period = "n"
            return rule, period
        return str(spec).strip(), "n"

    # ------------------------------------------------------------------
    # 1. DATA FETCHING
    # ------------------------------------------------------------------

    def _aggregate_balances(self, document_id_filter: str | None) -> tuple[dict, dict]:
        """Retourne (raw, balances) pour un filtre document donné (ou tout le dossier si None)."""
        query = (
            self.db.query(
                models.Account.code,
                func.coalesce(func.sum(models.EntryLine.debit), 0).label("debit"),
                func.coalesce(func.sum(models.EntryLine.credit), 0).label("credit"),
            )
            .join(models.EntryLine, models.EntryLine.account_id == models.Account.id)
            .join(models.Entry, models.Entry.id == models.EntryLine.entry_id)
            .join(models.Journal, models.Journal.id == models.Entry.journal_id)
            .filter(models.Account.company_id == self.company_id)
            .filter(models.Journal.company_id == self.company_id)
        )
        if document_id_filter:
            query = query.filter(models.Entry.document_id == document_id_filter)
        query = query.group_by(models.Account.code)

        raw: dict[str, tuple[float, float]] = {}
        balances: dict[str, float] = {}
        for row in query.all():
            d = float(row.debit)
            c = float(row.credit)
            code = self._normalize_code(row.code)
            if not code:
                continue
            raw[code] = (d, c)
            balances[code] = d - c
        return raw, balances

    def _load_annexe_data(self):
        import json

        annexe_record = (
            self.db.query(models.AnnexeData)
            .filter(models.AnnexeData.company_id == self.company_id)
            .first()
        )
        if annexe_record and annexe_record.data:
            try:
                self.annexe_data = json.loads(annexe_record.data)
            except Exception:
                pass

    def _fetch_balances(self):
        """
        Charge les soldes N (`document_id`) et optionnellement N-1 (`document_id_n1`).
        `self.balances` / `self.raw` restent l'exercice N pour compatibilité (pre-flight, contrôles).
        """
        self.raw, self.balances = self._aggregate_balances(self.document_id)
        if self.document_id_n1:
            self.raw_n1, self.balances_n1 = self._aggregate_balances(self.document_id_n1)
        else:
            self.raw_n1, self.balances_n1 = {}, {}

        self._load_annexe_data()

        self._log.append(
            f"_fetch_balances: N={len(self.balances)} accounts, "
            f"N-1={len(self.balances_n1)} accounts, "
            f"{len(self.annexe_data)} annexe variables."
        )

    # ------------------------------------------------------------------
    # 2. RULE ENGINE
    # ------------------------------------------------------------------

    def _resolve_rule(
        self,
        rule: str,
        cell_ref: str = "",
        strict_mode: bool = True,
        *,
        balances: dict[str, float] | None = None,
    ) -> tuple[float | str, list[str]]:
        """
        Parse a mapping rule string and compute the value.
        Supports variables extra-comptables via # (ex: "#nif", "#dirigeant_nom").
        Wrappers supported:
        - ABS() : Valeur absolue totale
        - SD()  : Solde Débiteur stricte (ignore les soldes créditeurs, renvoie positif)
        - SC()  : Solde Créditeur stricte (ignore les soldes débiteurs, renvoie positif)
        """
        rule = rule.strip()
        if not rule:
            return 0.0, []

        bal_map = balances if balances is not None else self.balances

        # Handle extra-accounting variable
        if rule.startswith("#") and "," not in rule:
            var_name = rule[1:]
            val = self.annexe_data.get(var_name, "")
            try:
                if str(val).replace('.', '', 1).isdigit():
                    return float(val), []
            except:
                pass
            return val, []

        # Handle Wrappers
        is_abs = False
        is_sd = False
        is_sc = False

        rule_upper = rule.upper()
        if rule_upper.startswith("ABS(") and rule_upper.endswith(")"):
            is_abs = True
            rule = rule[4:-1]
        elif rule_upper.startswith("SD(") and rule_upper.endswith(")"):
            is_sd = True
            rule = rule[3:-1]
        elif rule_upper.startswith("SC(") and rule_upper.endswith(")"):
            is_sc = True
            rule = rule[3:-1]

        total = 0.0
        patterns = [p.strip() for p in rule.split(",") if p.strip()]
        matched_codes_all: list[str] = []

        for pattern in patterns:
            multiplier = 1.0
            if pattern.startswith("-"):
                multiplier = -1.0
                pattern = pattern[1:].strip()

            component = 0.0
            matched_codes: list[str] = []
            if pattern.endswith("*"):
                prefix = pattern[:-1]
                # Un seul chiffre (classe incomplète) : trop large ; 2+ chiffres = racines SYSCOHADA usuelles (20, 41…).
                if len(prefix) <= 1:
                    msg = (
                        f"Règle trop large ({pattern}) sur {cell_ref or 'cellule inconnue'} ; "
                        "préférez au moins 2 chiffres significatifs (ex. 41* plutôt que 4*)."
                    )
                    if strict_mode:
                        self.mapping_errors.append(msg)
                    else:
                        self.mapping_warnings.append(msg)
                for code, bal in bal_map.items():
                    if code.startswith(prefix):
                        if is_sd and bal <= 0:
                            continue
                        if is_sc and bal >= 0:
                            continue
                        component += bal
                        matched_codes.append(code)
            else:
                bal = bal_map.get(pattern, 0.0)
                if not ((is_sd and bal <= 0) or (is_sc and bal >= 0)):
                    component += bal
                    if pattern in bal_map:
                        matched_codes.append(pattern)

            total += component * multiplier
            matched_codes_all.extend(matched_codes)

        # In strict mode, numeric rules must match at least one account.
        if strict_mode and self._is_numeric_expr(rule) and not matched_codes_all:
            self.mapping_errors.append(
                f"Aucun compte trouvé pour la règle '{rule}' sur {cell_ref or 'cellule inconnue'}."
            )

        if is_sd:
            result = max(0.0, total)
        elif is_sc:
            result = abs(min(0.0, total)) # SC returns positive value of credit balances
        elif is_abs:
            result = abs(total)
        else:
            result = total

        # Arrondi à l'entier près (0 décimale) comme exigé par la majorité des canevas OTR
        return float(round(result, 0)), sorted(set(matched_codes_all))

    # ------------------------------------------------------------------
    # 3. CELL WRITING (merged-cell safe, formula-safe)
    # ------------------------------------------------------------------

    def _parse_cell_addr(self, addr: str) -> tuple[int, int]:
        """Convert e.g. 'E13' → (row=13, col=5)."""
        col_letters = ""
        row_digits = ""
        for ch in addr:
            if ch.isalpha():
                col_letters += ch
            else:
                row_digits += ch
        return int(row_digits), column_index_from_string(col_letters)

    def inject_value(self, ws, cell_addr: str, value):
        """
        Write value to cell_addr in worksheet ws.
        If the cell belongs to a merged range, write to the top-left master.
        Always replaces any existing text/formula with the numeric value for targeted input cells.
        """
        target_row, target_col = self._parse_cell_addr(cell_addr)

        # Check if this cell is inside a merged range
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= target_row <= merged_range.max_row and
                    merged_range.min_col <= target_col <= merged_range.max_col):
                # Write to the top-left master cell of the merged range
                master = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                master.value = value
                self._log.append(f"  MERGED → {ws.title}!{cell_addr} → master "
                                  f"({get_column_letter(merged_range.min_col)}{merged_range.min_row}) = {value}")
                return

        # Normal (non-merged) cell
        cell = ws.cell(row=target_row, column=target_col)
        old_val = cell.value
        cell.value = value
        self._log.append(
            f"  WRITE  → {ws.title}!{cell_addr} = {value}"
            + (f"  [replaced formula: {str(old_val)[:30]}]" if isinstance(old_val, str) and old_val.startswith("=") else "")
        )

    # ------------------------------------------------------------------
    # 4. REPORT GENERATION & VALIDATION
    # ------------------------------------------------------------------

    def pre_flight_check(self) -> dict:
        """
        Vérification SQL en amont (Pre-flight Check) :
        1. Équilibre Bilan : SUM(Débit) = SUM(Crédit)
        2. Statistiques et identification du Compte 13 (Résultat Net)
        """
        self._fetch_balances()

        total_debit = sum(val[0] for val in self.raw.values())
        total_credit = sum(val[1] for val in self.raw.values())
        
        # Le résultat net se lit généralement sur le compte 13
        resultat_net = None
        for code, bal in self.balances.items():
            if code.startswith('13'):
                resultat_net = bal

        is_balanced = abs(total_debit - total_credit) < 1.0

        return {
            "total_debit": round(total_debit, 2),
            "total_credit": round(total_credit, 2),
            "difference": round(abs(total_debit - total_credit), 2),
            "is_balanced": is_balanced,
            "compte_13_present": resultat_net is not None,
            "resultat_net_13": round(resultat_net, 2) if resultat_net is not None else 0.0,
            "nb_comptes": len(self.balances)
        }

    def validate_totals(self, template_path: str, output_path: str) -> list[str]:
        """
        Vérification post-injection : s'assure qu'aucune cellule contenant originellement
        une formule de =SOMME (Total) n'a été écrasée par erreur.
        """
        wb_template = openpyxl.load_workbook(template_path, data_only=False, read_only=True)
        wb_output = openpyxl.load_workbook(output_path, data_only=False, read_only=True)
        
        errors = []
        for sheet_name in ["BILAN ACTIF", "BILAN PASSIF", "COMPTE DE RESULTAT", "Résultat fiscal"]:
            if sheet_name not in wb_template.sheetnames:
                continue
                
            ws_t = wb_template[sheet_name]
            if sheet_name not in wb_output.sheetnames:
                continue
            ws_o = wb_output[sheet_name]
            
            for row in ws_t.iter_rows(min_row=1, max_row=60, min_col=1, max_col=15):
                for cell in row:
                    val = str(cell.value).strip().upper() if cell.value is not None else ""
                    # On cherche les cellules qui contiennent =SUM ou =SOMME ou autres formules critiques
                    if val.startswith("=SOMME") or val.startswith("=SUM") or val.startswith("=SUBTOTAL") or val.startswith("=SOUS.TOTAL"):
                        out_val = str(ws_o[cell.coordinate].value).strip().upper() if ws_o[cell.coordinate].value is not None else ""
                        if not (out_val.startswith("=SOMME") or out_val.startswith("=SUM") or out_val.startswith("=SUBTOTAL") or out_val.startswith("=SOUS.TOTAL")):
                            errors.append(f"{sheet_name}!{cell.coordinate} (Formule écrasée : {out_val})")
        
        wb_template.close()
        wb_output.close()
        
        if errors:
            self._log.append(f"⚠️ AVERTISSEMENT VALIDATION: {len(errors)} formules de total/somme écrasees: {errors[:5]}")
            print(f"[ExcelInjector] WARN: {len(errors)} formules de TOTAL écrasées !")
        else:
            self._log.append("✅ VALIDATION: Toutes les formules de total (SOMME/SUM/SOUS.TOTAL) sont intactes.")
            print("[ExcelInjector] Validation OK : Formules intactes.")
            
        return errors

    def generate_report(
        self,
        template_path: str,
        output_path: str,
        mapping_config: dict,
        strict_mode: bool = True,
    ) -> str:
        """
        Copy the template, inject computed values, save to output_path.
        Returns output_path.
        """
        self._fetch_balances()

        if not self.balances:
            raise ValueError(
                "Aucun solde comptable trouvé pour cette société. "
                "Veuillez d'abord importer une balance générale."
            )

        # (do NOT use data_only=True — we want to preserve formulas)
        # keep_vba must only be enabled for macro-enabled templates (.xlsm/.xltm),
        # otherwise Excel may reject a generated .xlsx as invalid format.
        template_ext = os.path.splitext(template_path)[1].lower()
        keep_vba = template_ext in {".xlsm", ".xltm"}
        wb = openpyxl.load_workbook(template_path, keep_vba=keep_vba, data_only=False)

        injected = 0
        skipped_sheet = []
        skipped_zero = []

        for cell_ref, rule_spec in mapping_config.items():
            if "!" in cell_ref:
                sheet_name, cell_addr = cell_ref.split("!", 1)
            else:
                sheet_name = wb.active.title
                cell_addr = cell_ref

            if sheet_name not in wb.sheetnames:
                skipped_sheet.append(cell_ref)
                continue

            rule_str, period = self.parse_mapping_value(rule_spec)
            ws = wb[sheet_name]

            if period == "n1" and not self.document_id_n1:
                msg = (
                    f"Cellule {sheet_name}!{cell_addr} : période N-1 demandée mais aucun "
                    "document de balance N-1 (document_id_n1) n'a été fourni."
                )
                if strict_mode:
                    self.mapping_errors.append(msg)
                else:
                    self.mapping_warnings.append(msg)
                self.inject_value(ws, cell_addr, 0)
                trace_row: dict = {
                    "cell_ref": f"{sheet_name}!{cell_addr}",
                    "rule": rule_str,
                    "period": period,
                    "value": 0,
                    "matched_accounts_count": 0,
                    "matched_accounts_sample": [],
                }
                if isinstance(rule_spec, dict):
                    trace_row["mapping_entry"] = rule_spec
                self.trace.append(trace_row)
                continue

            bal_source = self.balances_n1 if period == "n1" else self.balances

            try:
                value, matched_codes = self._resolve_rule(
                    rule_str,
                    cell_ref=f"{sheet_name}!{cell_addr}",
                    strict_mode=strict_mode,
                    balances=bal_source,
                )
                # Write ALL mapped values including 0 (so template input zeros aren't kept as formulas)
                self.inject_value(ws, cell_addr, value)
                if value != 0:
                    injected += 1
                trace_row = {
                    "cell_ref": f"{sheet_name}!{cell_addr}",
                    "rule": rule_str,
                    "period": period,
                    "value": value,
                    "matched_accounts_count": len(matched_codes),
                    "matched_accounts_sample": matched_codes[:30],
                }
                if isinstance(rule_spec, dict):
                    trace_row["mapping_entry"] = rule_spec
                self.trace.append(trace_row)
            except Exception as exc:
                self._log.append(f"  ERROR  → {cell_ref} ({rule_spec}): {exc}")
                print(f"[ExcelInjector] WARN: {cell_ref} ({rule_spec}) → {exc}")

        if strict_mode and self.mapping_errors:
            raise ValueError(
                "Échec validation mapping strict. "
                + " | ".join(sorted(set(self.mapping_errors))[:8])
            )

        if skipped_sheet:
            msg = f"[ExcelInjector] Sheets absent du template: {set(skipped_sheet)}"
            self._log.append(msg)
            print(msg)

        print(f"[ExcelInjector] {injected} non-zero values injected into '{output_path}'")
        
        wb.save(output_path)

        # Traceability artifact: allows business/audit teams to verify each injected value.
        trace_path = f"{output_path}.trace.json"
        controls = compute_post_injection_controls(
            balances=self.balances,
            raw=self.raw,
            trace=self.trace,
            mapping_warnings=self.mapping_warnings,
            mapping_errors=self.mapping_errors,
        )
        with open(trace_path, "w", encoding="utf-8") as fp:
            json.dump(
                {
                    "company_id": self.company_id,
                    "document_id": self.document_id,
                    "document_id_n1": self.document_id_n1,
                    "generated_file": os.path.basename(output_path),
                    "injected_non_zero_count": injected,
                    "controls": controls,
                    "mapping_warnings": sorted(set(self.mapping_warnings)),
                    "mapping_errors": sorted(set(self.mapping_errors)),
                    "trace": self.trace,
                },
                fp,
                ensure_ascii=False,
                indent=2,
            )
        
        # Validation Post-injection
        self.validate_totals(template_path, output_path)
        
        return output_path

    @property
    def log(self) -> list[str]:
        return self._log


