"""
Génération d'une Balance Générale de test au format SYSCOHADA Révisé
Société fictive : ETAL COMMERCE SARL — Exercice 2025 — Lomé, Togo

Format : 6 colonnes
  Compte | Libellé | Débit Mouvements | Crédit Mouvements | Solde Débiteur | Solde Créditeur

Exécuter :
  cd backend
  python create_test_balance.py
→ Crée : test_balance_ETAL_COMMERCE_2025.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# DONNÉES — Balance équilibrée (Σ SD = Σ SC = 277 300 000 FCFA)
# ---------------------------------------------------------------------------
# Format : (code, libellé, débit_mvt, crédit_mvt, solde_débiteur, solde_créditeur)
# Les mouvements incluent les écritures d'ouverture + exercice courant
# Les soldes = position nette au 31/12/2025

LIGNES = [
    # ── CLASSE 1 : CAPITAUX ─────────────────────────────────────────────
    ("101",  "Capital social",                              0,           50_000_000,       0,      50_000_000),
    ("161",  "Emprunt bancaire BTCI",                       0,           20_000_000,       0,      20_000_000),

    # ── CLASSE 2 : IMMOBILISATIONS ──────────────────────────────────────
    ("241",  "Matériel et outillage industriel",       12_000_000,              0,  12_000_000,             0),
    ("245",  "Matériel de transport",                   8_000_000,              0,   8_000_000,             0),
    ("284",  "Amort. matériel, mobilier et agencement",         0,       1_000_000,          0,     1_000_000),
    ("285",  "Amort. matériel de transport",                    0,       1_000_000,          0,     1_000_000),

    # ── CLASSE 3 : STOCKS ───────────────────────────────────────────────
    ("311",  "Stocks de marchandises",                 90_000_000,      70_000_000,  20_000_000,            0),

    # ── CLASSE 4 : TIERS ────────────────────────────────────────────────
    ("401",  "Fournisseurs — Achats courants",          10_000_000,      98_000_000,          0,    88_000_000),
    ("411",  "Clients — Ventes courantes",             135_700_000,     120_700_000,  15_000_000,            0),
    ("421",  "Personnel — Rémunérations dues",          10_000_000,      12_000_000,          0,     2_000_000),

    # ── CLASSE 5 : TRÉSORERIE ───────────────────────────────────────────
    ("521",  "Banque BTCI Lomé — Compte courant",      174_300_000,      45_800_000, 128_500_000,            0),

    # ── CLASSE 6 : CHARGES D'EXPLOITATION ───────────────────────────────
    ("601",  "Achats de marchandises",                  70_000_000,              0,  70_000_000,            0),
    ("611",  "Transports sur achats",                    2_500_000,              0,   2_500_000,            0),
    ("621",  "Redevances crédit-bail et locations",      4_800_000,              0,   4_800_000,            0),
    ("631",  "Honoraires et frais d'actes",              1_200_000,              0,   1_200_000,            0),
    ("641",  "Impôts, taxes et versements assimilés",      500_000,              0,     500_000,            0),
    ("657",  "Amendes, pénalités et majoration",           150_000,              0,     150_000,            0),
    ("661",  "Salaires et traitements bruts",            12_000_000,              0,  12_000_000,            0),
    ("671",  "Charges d'intérêts sur emprunts",            650_000,              0,     650_000,            0),
    ("681",  "Dotations aux amortissements",              2_000_000,              0,   2_000_000,            0),

    # ── CLASSE 7 : PRODUITS D'EXPLOITATION ──────────────────────────────
    ("701",  "Ventes de marchandises",                           0,     115_000_000,          0,   115_000_000),
    ("771",  "Intérêts et produits assimilés",                   0,         300_000,          0,       300_000),
]

# Vérification automatique de l'équilibre
total_sd = sum(l[4] for l in LIGNES)
total_sc = sum(l[5] for l in LIGNES)
assert total_sd == total_sc, f"ERREUR : La balance n'est PAS équilibrée ! SD={total_sd:,} ≠ SC={total_sc:,}"
print(f"✅ Balance équilibrée : Σ SD = Σ SC = {total_sd:,.0f} FCFA")

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", fgColor="1B4F8A")   # Bleu OTR
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

CL1_FILL      = PatternFill("solid", fgColor="D6E4F7")   # Bleu clair — Capitaux
CL2_FILL      = PatternFill("solid", fgColor="D5F5E3")   # Vert clair — Immob.
CL3_FILL      = PatternFill("solid", fgColor="FEF9E7")   # Jaune clair — Stocks
CL4_FILL      = PatternFill("solid", fgColor="FDEDEC")   # Rose clair  — Tiers
CL5_FILL      = PatternFill("solid", fgColor="EBF5FB")   # Cyan clair  — Tréso
CL6_FILL      = PatternFill("solid", fgColor="FDF2E9")   # Orange clair — Charges
CL7_FILL      = PatternFill("solid", fgColor="E9F7EF")   # Vert clair  — Produits

TOTAL_FILL    = PatternFill("solid", fgColor="2C3E50")
TOTAL_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

DEF_FONT      = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)
NUM_FORMAT    = '#,##0'

thin = Side(style="thin", color="AAAAAA")
thick = Side(style="medium", color="1B4F8A")
BORDER_STD = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)

CLASS_FILL = {
    "1": CL1_FILL, "2": CL2_FILL, "3": CL3_FILL,
    "4": CL4_FILL, "5": CL5_FILL, "6": CL6_FILL, "7": CL7_FILL,
}

CLASS_LABELS = {
    "1": "CLASSE 1 — RESSOURCES DURABLES",
    "2": "CLASSE 2 — VALEURS IMMOBILISÉES",
    "3": "CLASSE 3 — STOCKS",
    "4": "CLASSE 4 — TIERS",
    "5": "CLASSE 5 — TRÉSORERIE",
    "6": "CLASSE 6 — CHARGES",
    "7": "CLASSE 7 — PRODUITS",
}

# ---------------------------------------------------------------------------
# CRÉATION DU CLASSEUR
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Balance Générale"

# ---- Titre ----
ws.merge_cells("A1:F1")
ws["A1"] = "ETAL COMMERCE SARL — BALANCE GÉNÉRALE AU 31/12/2025"
ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="1B4F8A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:F2")
ws["A2"] = "RC Togo 2020B12345 — NIF 1234567890 — Secteur : Commerce Général — Lomé"
ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="555555")
ws["A2"].alignment = Alignment(horizontal="center")

ws.merge_cells("A3:F3")  # spacer

# ---- En-têtes colonnes ----
HEADERS = ["Compte", "Libellé", "Débit Mouvements", "Crédit Mouvements", "Solde Débiteur", "Solde Créditeur"]
HDR_ROW = 4
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=HDR_ROW, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_STD
ws.row_dimensions[HDR_ROW].height = 32

# ---- Largeur colonnes ----
COL_WIDTHS = [12, 42, 20, 20, 20, 20]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---- Lignes de données ----
data_row = HDR_ROW + 1
seen_classes = set()

for code, libelle, d_mvt, c_mvt, sd, sc in LIGNES:
    cls = code[0]
    fill = CLASS_FILL.get(cls, PatternFill())

    # Séparateur de classe
    if cls not in seen_classes:
        seen_classes.add(cls)
        ws.merge_cells(f"A{data_row}:F{data_row}")
        sep_cell = ws[f"A{data_row}"]
        sep_cell.value = CLASS_LABELS.get(cls, f"CLASSE {cls}")
        sep_cell.font = Font(name="Calibri", bold=True, size=9, color="1B4F8A")
        sep_cell.fill = PatternFill("solid", fgColor="EAF2FB")
        sep_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sep_cell.border = Border(bottom=Side(style="thin", color="1B4F8A"))
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # Ligne de compte
    row_data = [code, libelle, d_mvt if d_mvt else None, c_mvt if c_mvt else None,
                sd if sd else None, sc if sc else None]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=data_row, column=col_idx, value=val)
        cell.font = DEF_FONT
        cell.fill = fill
        cell.border = BORDER_STD
        if col_idx == 1:
            cell.font = Font(name="Courier New", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            cell.number_format = NUM_FORMAT
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[data_row].height = 17
    data_row += 1

# ---- Ligne TOTAL ----
data_row += 1
ws.merge_cells(f"A{data_row}:B{data_row}")
total_label = ws[f"A{data_row}"]
total_label.value = "TOTAL GÉNÉRAL"
total_label.font = TOTAL_FONT
total_label.fill = TOTAL_FILL
total_label.alignment = Alignment(horizontal="center", vertical="center")
total_label.border = BORDER_STD

for col_offset, val in enumerate([
    sum(l[2] for l in LIGNES),
    sum(l[3] for l in LIGNES),
    total_sd,
    total_sc,
], 3):
    cell = ws.cell(row=data_row, column=col_offset, value=val)
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.number_format = NUM_FORMAT
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = BORDER_STD
ws.row_dimensions[data_row].height = 22

# ---- Équilibre check ----
data_row += 2
ws.merge_cells(f"A{data_row}:F{data_row}")
check_cell = ws[f"A{data_row}"]
check_cell.value = (
    f"✅  Balance ÉQUILIBRÉE — Σ Soldes Débiteurs = Σ Soldes Créditeurs = {total_sd:,.0f} FCFA  |  "
    f"Résultat d'Exercice implicite (Produits - Charges) = {(sum(l[5] for l in LIGNES if l[0].startswith('7')) - sum(l[4] for l in LIGNES if l[0].startswith('6'))):,.0f} FCFA"
)
check_cell.font = Font(name="Calibri", bold=True, size=9, color="1A5276")
check_cell.alignment = Alignment(horizontal="center")

# ---- Freeze panes ----
ws.freeze_panes = ws[f"A{HDR_ROW + 1}"]

# ---------------------------------------------------------------------------
# SAUVEGARDE
# ---------------------------------------------------------------------------
OUTPUT_NAME = "test_balance_ETAL_COMMERCE_2025.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), OUTPUT_NAME)
wb.save(OUTPUT_PATH)
print(f"\n📁 Fichier généré : {OUTPUT_PATH}")
print(f"\n📊 Résumé de la balance :")
print(f"   Nombre de comptes  : {len(LIGNES)}")
print(f"   Σ Soldes Débiteurs : {total_sd:>20,.0f} FCFA")
print(f"   Σ Soldes Créditeurs: {total_sc:>20,.0f} FCFA")
profit = sum(l[5] for l in LIGNES if l[0].startswith('7')) - \
         sum(l[4] for l in LIGNES if l[0].startswith('6'))
print(f"   Résultat Exercice  : {profit:>20,.0f} FCFA  (Bénéfice)" if profit >= 0 else
      f"   Résultat Exercice  : {profit:>20,.0f} FCFA  (Perte)")
print(f"\n✅ Prêt à importer dans Auditia via : Import > Importer une Balance")
