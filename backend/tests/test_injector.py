import json
import os
import tempfile

import openpyxl
import pytest

from app.services.injector import ExcelInjector
from app.services.injector_controls import compute_post_injection_controls


def test_normalize_code():
    assert ExcelInjector._normalize_code("411.0") == "4110"
    assert ExcelInjector._normalize_code("401-001") == "401001"
    assert ExcelInjector._normalize_code("") == ""


def test_resolve_rule_strict_rejects_single_digit_wildcard(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    inj._fetch_balances()
    inj.mapping_errors.clear()
    inj._resolve_rule("4*", cell_ref="Feuil1!E1", strict_mode=True)
    assert any("trop large" in e.lower() or "large" in e.lower() for e in inj.mapping_errors)


def test_resolve_rule_non_strict_warns_single_digit_wildcard(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    inj._fetch_balances()
    inj.mapping_warnings.clear()
    inj.mapping_errors.clear()
    inj._resolve_rule("4*", cell_ref="Feuil1!E1", strict_mode=False)
    assert any("large" in w.lower() for w in inj.mapping_warnings)
    assert not inj.mapping_errors


def test_parse_mapping_value():
    assert ExcelInjector.parse_mapping_value("411*") == ("411*", "n")
    assert ExcelInjector.parse_mapping_value({"rule": "101", "period": "n1"}) == ("101", "n1")


def test_resolve_rule_strict_no_match(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    inj._fetch_balances()
    inj.mapping_errors.clear()
    inj._resolve_rule("999", cell_ref="Feuil1!A1", strict_mode=True)
    assert any("Aucun compte trouvé" in e for e in inj.mapping_errors)


def test_resolve_rule_prefix_three_digits_ok(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    inj._fetch_balances()
    inj.mapping_errors.clear()
    val, codes = inj._resolve_rule("411*", cell_ref="Feuil1!A1", strict_mode=True)
    assert not inj.mapping_errors
    assert "411" in codes
    assert val == 1000.0


def test_signed_wrapper_for_compte_13_semantics(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    inj._fetch_balances()
    inj.balances["13"] = -3000.0  # crediteur
    val, _ = inj._resolve_rule("SIGNED(13)", cell_ref="Passif!X1", strict_mode=True)
    assert val == 3000.0
    inj.balances["13"] = 2500.0  # debiteur
    val2, _ = inj._resolve_rule("SIGNED(13)", cell_ref="Passif!X2", strict_mode=True)
    assert val2 == -2500.0


def test_pre_flight_balanced(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    pf = inj.pre_flight_check()
    assert pf["is_balanced"] is True
    assert pf["nb_comptes"] == 2


def test_compute_post_injection_controls(db_session, seed_company_balanced):
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    inj._fetch_balances()
    rep = compute_post_injection_controls(
        balances=inj.balances,
        raw=inj.raw,
        trace=[
            {
                "cell_ref": "Bilan!E10",
                "rule": "101",
                "value": -1000.0,
                "matched_accounts_count": 1,
            }
        ],
        mapping_warnings=[],
        mapping_errors=[],
    )
    assert rep["general_ledger"]["balanced"] is True
    assert rep["net_balances"]["closure_ok"] is True
    assert rep["compte_13"]["present"] is False


def test_generate_report_minimal_template(db_session, seed_company_balanced):
    cid = seed_company_balanced["company_id"]
    inj = ExcelInjector(db_session, cid)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = 0
    path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    out = path.replace(".xlsx", "_out.xlsx")
    wb.save(path)
    try:
        inj.generate_report(
            template_path=path,
            output_path=out,
            mapping_config={"TestSheet!A1": "411*"},
            strict_mode=True,
        )
        wbo = openpyxl.load_workbook(out, data_only=True)
        assert wbo["TestSheet"]["A1"].value == 1000
        trace_path = f"{out}.trace.json"
        with open(trace_path, "r", encoding="utf-8") as fp:
            payload = json.load(fp)
        assert "controls" in payload
        assert payload["controls"]["general_ledger"]["balanced"] is True
    finally:
        for p in (path, out, f"{out}.trace.json"):
            if os.path.exists(p):
                os.unlink(p)


def test_generate_report_n_n1_two_documents(db_session, seed_company_nn1):
    info = seed_company_nn1
    inj = ExcelInjector(
        db_session,
        info["company_id"],
        document_id=info["document_n"],
        document_id_n1=info["document_n1"],
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bilan"
    path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    out = path.replace(".xlsx", "_nn1.xlsx")
    wb.save(path)
    try:
        inj.generate_report(
            template_path=path,
            output_path=out,
            mapping_config={
                "Bilan!A1": "411*",
                "Bilan!B1": {"rule": "411*", "period": "n1"},
            },
            strict_mode=True,
        )
        wbo = openpyxl.load_workbook(out, data_only=True)
        assert wbo["Bilan"]["A1"].value == 1000
        assert wbo["Bilan"]["B1"].value == 400
        with open(f"{out}.trace.json", "r", encoding="utf-8") as fp:
            payload = json.load(fp)
        assert payload.get("document_id_n1") == info["document_n1"]
        periods = {t["cell_ref"]: t.get("period") for t in payload["trace"]}
        assert periods.get("Bilan!B1") == "n1"
    finally:
        for p in (path, out, f"{out}.trace.json"):
            if os.path.exists(p):
                os.unlink(p)


def test_generate_report_strict_raises_on_bad_rule(db_session, seed_company_balanced):
    path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    wb = openpyxl.Workbook()
    wb.active.title = "S"
    wb.save(path)
    out = path.replace(".xlsx", "_out2.xlsx")
    inj = ExcelInjector(db_session, seed_company_balanced["company_id"])
    try:
        with pytest.raises(ValueError, match="validation mapping strict"):
            inj.generate_report(
                template_path=path,
                output_path=out,
                mapping_config={"S!A1": "999"},
                strict_mode=True,
            )
    finally:
        for p in (path,):
            if os.path.exists(p):
                os.unlink(p)
        if os.path.exists(out):
            os.unlink(out)
