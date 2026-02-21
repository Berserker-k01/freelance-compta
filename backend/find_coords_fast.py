import openpyxl

def find_text_fast(file_path, keywords):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    results = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Only check first 100 rows
        for row_idx, row in enumerate(ws.iter_rows(max_row=100)):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for kw in keywords:
                        if kw.lower() in cell.value.lower():
                            if kw not in results: results[kw] = []
                            results[kw].append(f"{sheet_name}!{cell.coordinate}: {cell.value}")
    return results

if __name__ == "__main__":
    template = "c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia/backend/templates/syscohada_template.xlsx"
    keywords = ["Matériel", "Capital", "Chiffre d'affaires", "Stocks", "Clients", "Banque", "Fournisseurs"]
    print(find_text_fast(template, keywords))
