import openpyxl

def extract_layout(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = ["BILAN ACTIF", "BILAN PASSIF", "COMPTE DE RESULTAT"]
    for s_name in sheets:
        if s_name not in wb.sheetnames: continue
        print(f"\n--- {s_name} ---")
        ws = wb[s_name]
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
            row_data = [cell.value for cell in row]
            if any(row_data):
                print(f"Row {row[0].row}: {row_data}")

if __name__ == "__main__":
    template = "c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia/backend/templates/syscohada_template.xlsx"
    extract_layout(template)
