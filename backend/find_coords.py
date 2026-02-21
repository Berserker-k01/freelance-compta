import openpyxl
import sys

def find_text_in_excel(file_path, search_texts):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        results = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for text in search_texts:
                            if text.lower() in cell.value.lower():
                                if text not in results: results[text] = []
                                results[text].append(f"{sheet_name}!{cell.coordinate}: {cell.value}")
        return results
    except Exception as e:
        return str(e)

if __name__ == "__main__":
    template = "c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia/backend/templates/syscohada_template.xlsx"
    keywords = ["Matériel", "Capital", "Chiffre d'affaires", "Stocks", "Clients", "Banque", "Fournisseurs"]
    res = find_text_in_excel(template, keywords)
    print(res)
