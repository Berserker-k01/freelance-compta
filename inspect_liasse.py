import openpyxl

# File with accent in name - easier via script
import glob, os

# Find the file
base = r"c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia"
matches = glob.glob(os.path.join(base, "Liasse*.xlsx"))
print("Found files:", matches)

if not matches:
    print("ERROR: no liasse file found")
    exit(1)

PATH = matches[0]
print(f"Inspecting: {PATH}\n")

wb = openpyxl.load_workbook(PATH, data_only=True)

print("=== SHEETS ===")
for s in wb.sheetnames:
    print(f"  '{s}'")
print()

TARGET_SHEETS = ["BILAN ACTIF", "BILAN PASSIF", "COMPTE DE RESULTAT", "Résultat fiscal"]

for sname in TARGET_SHEETS:
    if sname not in wb.sheetnames:
        print(f"❌ SHEET MISSING: '{sname}'")
        continue

    ws = wb[sname]
    print(f"\n{'='*55}")
    print(f"  {sname}")
    print(f"{'='*55}")

    found = 0
    for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=12):
        for cell in row:
            val = cell.value
            if val is None:
                continue
            sv = str(val).strip()
            if sv in ("", "0", "0.0"):
                continue
            try:
                fv = float(sv)
                # Only show numeric values > 0 (our injected data)
                if fv != 0:
                    print(f"  {cell.coordinate:6s} = {fv:>15,.2f}")
                    found += 1
            except (ValueError, TypeError):
                # Text cell - skip unless it's in column D+ (where numbers should be)
                if cell.column >= 4:
                    print(f"  {cell.coordinate:6s} = '{sv[:50]}'")
                    found += 1

    if found == 0:
        print(f"  ⚠️  TOUTES LES CELLULES NUMÉRIQUES SONT VIDES (lignes 1-60, cols A-L)")
        # Show what's there anyway (any non-null)
        print("  Quelques valeurs présentes (texte inclus) :")
        cnt = 0
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12):
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    print(f"    {cell.coordinate}: {str(cell.value)[:60]}")
                    cnt += 1
                    if cnt > 15:
                        break
            if cnt > 15:
                break

print("\n=== DONE ===")
