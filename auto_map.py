"""
Detecte la structure des colonnes dans BILAN ACTIF, BILAN PASSIF, 
COMPTE DE RESULTAT, Resultat fiscal.
Ecrit les resultats dans un fichier JSON pour eviter les problemes d'encodage.
"""
import openpyxl
import json

TEMPLATE = r"backend/templates/syscohada_template.xlsx"
OUTPUT = "template_structure.json"

print("Chargement du template (patience)...")
wb = openpyxl.load_workbook(TEMPLATE, data_only=False, read_only=True, keep_vba=False)

KEY_SHEETS = {
    "BILAN ACTIF": {"rows": (1, 100), "max_col": 20},
    "BILAN PASSIF": {"rows": (1, 100), "max_col": 20},
    "COMPTE DE RESULTAT": {"rows": (1, 60), "max_col": 20},
}

# Try both names for Resultat fiscal
for name in ["Resultat fiscal", "R\u00e9sultat fiscal", "RESULTAT FISCAL"]:
    if name in wb.sheetnames:
        KEY_SHEETS[name] = {"rows": (1, 50), "max_col": 15}
        break

result = {}

for sname, cfg in KEY_SHEETS.items():
    if sname not in wb.sheetnames:
        print(f"MISSING: {sname}")
        continue

    print(f"Scanning: {sname} ...")
    ws = wb[sname]
    r_min, r_max = cfg["rows"]
    max_col = cfg["max_col"]

    # Count types per column
    col_counts = {}
    sample_cells = []  # (row, col, letter, value)

    for row in ws.iter_rows(min_row=r_min, max_row=r_max,
                             min_col=1, max_col=max_col):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            col = cell.column
            sv = str(v).strip()
            if not sv:
                continue

            if sv.startswith("="):
                col_counts.setdefault(col, {"formula": 0, "number": 0, "text": 0})
                col_counts[col]["formula"] += 1
            else:
                try:
                    fv = float(sv)
                    col_counts.setdefault(col, {"formula": 0, "number": 0, "text": 0})
                    col_counts[col]["number"] += 1
                    col_letter = chr(64 + col) if col <= 26 else f"col{col}"
                    sample_cells.append({
                        "coord": f"{col_letter}{cell.row}",
                        "col": col_letter,
                        "row": cell.row,
                        "value": fv
                    })
                except ValueError:
                    col_counts.setdefault(col, {"formula": 0, "number": 0, "text": 0})
                    col_counts[col]["text"] += 1

    # Determine dominant type per column
    col_summary = {}
    for col, counts in sorted(col_counts.items()):
        dominant = max(counts, key=counts.get)
        letter = chr(64 + col) if col <= 26 else f"col{col}"
        col_summary[letter] = {
            "dominant": dominant,
            "counts": counts
        }

    result[sname] = {
        "col_summary": col_summary,
        "numeric_values_found": sample_cells[:50]  # first 50
    }

wb.close()

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(result, f, indent=2, ensure_ascii=True)

print(f"\nResultats sauvegardes dans: {OUTPUT}")

# Print summary to stdout (ASCII only)
print("\n=== SUMMARY ===")
for sname, info in result.items():
    numeric_cols = [col for col, d in info["col_summary"].items() if d["dominant"] == "number"]
    formula_cols = [col for col, d in info["col_summary"].items() if d["dominant"] == "formula"]
    n_vals = len(info["numeric_values_found"])
    print(f"  {sname[:35]:<35s} | numeric_cols={numeric_cols} | formula_cols={formula_cols} | values={n_vals}")

print("=== DONE ===")
