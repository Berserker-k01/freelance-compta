"""
Inspecte rapidement juste BILAN ACTIF dans le template et la liasse
"""
import openpyxl
import glob

TEMPLATE = r"c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia/backend/templates/syscohada_template.xlsx"
LIASSE   = glob.glob(r"c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia/Liasse*.xlsx")[0]

print(f"Liasse: {LIASSE}")

# TEMPLATE - load ONLY the BILAN ACTIF sheet
wb_t = openpyxl.load_workbook(TEMPLATE, data_only=False, read_only=True)
if "BILAN ACTIF" in wb_t.sheetnames:
    ws = wb_t["BILAN ACTIF"]
    print("\n=== TEMPLATE — BILAN ACTIF (lignes 1-35) ===")
    for row in ws.iter_rows(min_row=1, max_row=35, min_col=1, max_col=10, values_only=False):
        for cell in row:
            v = cell.value
            if v is None: continue
            sv = str(v).strip()
            if not sv: continue
            # show everything
            print(f"  {cell.coordinate:7s}: {sv[:70]}")
else:
    print("BILAN ACTIF not found in template")
    print("Available sheets:", wb_t.sheetnames[:15])
wb_t.close()

# LIASSE - check what's in BILAN ACTIF
wb_l = openpyxl.load_workbook(LIASSE, data_only=False, read_only=True)
if "BILAN ACTIF" in wb_l.sheetnames:
    ws = wb_l["BILAN ACTIF"]
    print("\n=== LIASSE GÉNÉRÉE — BILAN ACTIF (lignes 1-35) ===")
    for row in ws.iter_rows(min_row=1, max_row=35, min_col=1, max_col=10, values_only=False):
        for cell in row:
            v = cell.value
            if v is None: continue
            sv = str(v).strip()
            if not sv: continue
            if not sv.startswith("="):  # Skip formulas in output
                try:
                    fv = float(sv)
                    if abs(fv) > 100:
                        print(f"  {cell.coordinate:7s}: {fv:>15,.2f}")
                except Exception:
                    if len(sv) > 1:
                        print(f"  {cell.coordinate:7s}: {sv[:60]}")
wb_l.close()

print("\n=== FIN ===")
