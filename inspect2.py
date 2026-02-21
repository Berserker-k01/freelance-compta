"""
Inspection minimaliste du template - colonnes de BILAN ACTIF
"""
import openpyxl
import glob
import sys

TEMPLATE = r"c:/Users/ROG FLOW/OneDrive/Documents/GitHub/Auditia/backend/templates/syscohada_template.xlsx"

try:
    # Load in read_only to be fast
    wb = openpyxl.load_workbook(TEMPLATE, data_only=True, read_only=True)
    
    # Show available sheets matching our targets
    targets = [s for s in wb.sheetnames if any(
        k in s.upper() for k in ["BILAN", "RESULTAT", "TFT", "COMPTE"]
    )]
    print("TARGET SHEETS FOUND:")
    for t in targets:
        print(f"  '{t}'")
    print()
    
    # Just check BILAN ACTIF structure rows 1-40
    sname = "BILAN ACTIF"
    if sname in wb.sheetnames:
        ws = wb[sname]
        print(f"BILAN ACTIF - all non-empty cells rows 1-40:")
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=10):
            for cell in row:
                if cell.value is not None:
                    v = str(cell.value)
                    if v.strip():
                        col = chr(64 + cell.column)
                        # Only show data cells (ignore long text)
                        if len(v) < 80:
                            print(f"  {col}{cell.row}: {v[:60]!r}")
    
    wb.close()

except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
